
import clr, sys, re, math, webbrowser
from pathlib import Path
from datetime import datetime

if not hasattr(clr, '_available_namespaces'):
    clr._available_namespaces = {}

import openpyxl
from openpyxl.styles import PatternFill, Font as XFont

clr.AddReference("RevitAPI")
from Autodesk.Revit.DB import *

doc = __revit__.ActiveUIDocument.Document #type:ignore
app = __revit__.Application #type:ignore

EXCEL_PATH = r'C:\Users\34655\OneDrive - RAEN Digital Tools SL\01_Proyectos\04_Editeca\01_Curso Revit IA\01_WIP\Caso 4_Control de calidad\Editeca_Control de Calidad.xlsx'
OUTPUT_DIR = str(Path(EXCEL_PATH).parent)
TODAY = datetime.now().strftime('%Y%m%d_%H%M%S')
DATE_STR = datetime.now().strftime('%d/%m/%Y %H:%M')

def ft_to_mm(v): return v * 304.8
def get_pval(p):
    try:
        v = p.AsString(); return v if v is not None else (p.AsValueString() or '')
    except: return ''
def eid_val(eid):
    try: return int(eid.Value)
    except: return int(eid.IntegerValue)

SHEET_BIP = {
    'Sheet Number': BuiltInParameter.SHEET_NUMBER, 'Sheet Name': BuiltInParameter.SHEET_NAME,
    'Drawn By': BuiltInParameter.SHEET_DRAWN_BY, 'Checked By': BuiltInParameter.SHEET_CHECKED_BY,
    'Approved By': BuiltInParameter.SHEET_APPROVED_BY, 'Sheet Issue Date': BuiltInParameter.SHEET_ISSUE_DATE,
    'Scale': BuiltInParameter.SHEET_SCALE,
}
def get_sheet_pval(sheet, pname):
    p = None
    bip = SHEET_BIP.get(pname)
    if bip:
        try: p = sheet.get_Parameter(bip)
        except: pass
    if p is None: p = sheet.LookupParameter(pname)
    return get_pval(p) if p else None

BIC_MAP = {
    'OST_Walls': BuiltInCategory.OST_Walls, 'OST_Floors': BuiltInCategory.OST_Floors,
    'OST_Roofs': BuiltInCategory.OST_Roofs, 'OST_Ceilings': BuiltInCategory.OST_Ceilings,
    'OST_Doors': BuiltInCategory.OST_Doors, 'OST_Windows': BuiltInCategory.OST_Windows,
    'OST_Stairs': BuiltInCategory.OST_Stairs, 'OST_Ramps': BuiltInCategory.OST_Ramps,
    'OST_StairsRailing': BuiltInCategory.OST_StairsRailing, 'OST_Casework': BuiltInCategory.OST_Casework,
    'OST_Furniture': BuiltInCategory.OST_Furniture, 'OST_CurtainWallPanels': BuiltInCategory.OST_CurtainWallPanels,
    'OST_CurtainWallMullions': BuiltInCategory.OST_CurtainWallMullions, 'OST_Rooms': BuiltInCategory.OST_Rooms,
    'OST_StructuralColumns': BuiltInCategory.OST_StructuralColumns, 'OST_StructuralFraming': BuiltInCategory.OST_StructuralFraming,
    'OST_StructuralFoundation': BuiltInCategory.OST_StructuralFoundation, 'OST_Rebar': BuiltInCategory.OST_Rebar,
    'OST_DuctCurves': BuiltInCategory.OST_DuctCurves, 'OST_FlexDuctCurves': BuiltInCategory.OST_FlexDuctCurves,
    'OST_DuctFitting': BuiltInCategory.OST_DuctFitting, 'OST_DuctAccessory': BuiltInCategory.OST_DuctAccessory,
    'OST_DuctTerminal': BuiltInCategory.OST_DuctTerminal, 'OST_PipeCurves': BuiltInCategory.OST_PipeCurves,
    'OST_FlexPipeCurves': BuiltInCategory.OST_FlexPipeCurves, 'OST_PipeFitting': BuiltInCategory.OST_PipeFitting,
    'OST_PipeAccessory': BuiltInCategory.OST_PipeAccessory, 'OST_MechanicalEquipment': BuiltInCategory.OST_MechanicalEquipment,
    'OST_ElectricalEquipment': BuiltInCategory.OST_ElectricalEquipment, 'OST_LightingFixtures': BuiltInCategory.OST_LightingFixtures,
    'OST_PlumbingFixtures': BuiltInCategory.OST_PlumbingFixtures, 'OST_ElectricalFixtures': BuiltInCategory.OST_ElectricalFixtures,
    'OST_Sprinklers': BuiltInCategory.OST_Sprinklers, 'OST_CableTray': BuiltInCategory.OST_CableTray,
    'OST_Topography': BuiltInCategory.OST_Topography, 'OST_Site': BuiltInCategory.OST_Site,
}

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
def srows(n): return [list(r) for r in wb[n].iter_rows(values_only=True) if any(x is not None for x in r)]

fichero_cfg  = {r[0]: r[1] for r in srows('Fichero')[1:]  if r[0]}
modelo_cfg   = {r[0]: r[1] for r in srows('Modelo')[1:]   if r[0]}
coord_cfg    = {r[0]: r[1] for r in srows('Coordenadas')[1:] if r[0]}
disciplinas  = {r[0]: r[1] for r in srows('Disciplinas')[1:] if r[0]}
planos_cfg   = [r for r in srows('Planos')[1:]   if r[0]]
proyecto_cfg = [r for r in srows('Proyecto')[1:] if r[0]]
clasificaciones = {}
for r in srows('Clasificaciones')[1:]:
    if r[2]:
        clasificaciones.setdefault(r[2], [])
        if r[3] and r[4]: clasificaciones[r[2]].append((str(r[3]), str(r[4])))
cod_to_nombre = {cod: nom for items in clasificaciones.values() for nom, cod in items}
mat_rows = srows('Matriz'); param_names = mat_rows[0][3:]
matrix = {}; cat_info = {}
for r in mat_rows[1:]:
    if r[2]:
        bic_str = r[2]; cat_info[bic_str] = (r[0], r[1]); matrix[bic_str] = {}
        for i, pn in enumerate(param_names):
            if pn and len(r) > 3+i and r[3+i]: matrix[bic_str][pn] = r[3+i]

results = {}
def add(sec, check, status, detalle, eid=None):
    results.setdefault(sec, [])
    results[sec].append({'Check': check, 'Estado': status, 'Detalle': str(detalle), 'ElementId': str(eid or '')})

def nom_errors(name, disc_code, cat_name, valid_codes, clasificacion_param_val=None):
    parts = name.split('_')
    errs = []
    if len(parts) < 4:
        if ' - ' in name: sep_det = '" - " en lugar de "_"'
        elif '-' in name: sep_det = '"-" en lugar de "_"'
        elif ' ' in name: sep_det = 'espacio en lugar de "_"'
        else: sep_det = 'no se detecta separador "_"'
        desc_limpia = re.sub(r'[\-]+', ' ', name).strip()[:25]
        cod_ejemplo = valid_codes[0] if valid_codes else 'COD'
        ejemplo = f'{disc_code}_{cat_name}_{cod_ejemplo}_{desc_limpia}'
        errs.append(f'Solo {len(parts)} parte(s): usa {sep_det}. Formato: DISC_Categoria_Clasificacion_Descripcion. Ejemplo: "{ejemplo}"')
    else:
        if parts[0] != disc_code:
            errs.append(f'Disciplina incorrecta: "{parts[0]}" → debe ser "{disc_code}"')
        if parts[1] != cat_name:
            errs.append(f'Categoria incorrecta: "{parts[1]}" → debe ser "{cat_name}"')
        if valid_codes and parts[2] not in valid_codes:
            errs.append(f'Clasificacion "{parts[2]}" no valida para {cat_name}. Valores aceptados: {", ".join(valid_codes)}')
        elif valid_codes and not errs and clasificacion_param_val:
            cv = clasificacion_param_val; exp_n = cod_to_nombre.get(parts[2], '')
            if cv and exp_n and cv != exp_n:
                errs.append(f'Inconsistencia: param Editeca_3D_Clasificacion="{cv}" pero codigo en nombre "{parts[2]}"="{exp_n}".')
    return errs

# A: FICHERO
title = doc.Title; path = doc.PathName
parts = title.replace('.rvt','').split('_'); valid_discs = list(disciplinas.values()) + ['GEN']
if len(parts) < 3: add('Fichero','Nombre','ERROR',f'"{title}" — necesita 3 partes separadas por "_": Autor_Disciplina_Descripcion')
elif parts[1] not in valid_discs: add('Fichero','Nombre','ERROR',f'Disciplina "{parts[1]}" no valida. Validas: {", ".join(valid_discs)}')
else: add('Fichero','Nombre','OK',title)
if path:
    try:
        mb = Path(path).stat().st_size/1048576; mx = float(fichero_cfg.get('Tamanio_Max_MB',300))
        add('Fichero','Tamanio','OK' if mb<=mx else 'ERROR',f'{mb:.1f} MB (max {mx:.0f} MB)')
    except Exception as e: add('Fichero','Tamanio','WARNING',str(e))
else: add('Fichero','Tamanio','WARNING','Fichero no guardado en disco todavia')
ver = int(app.VersionNumber); min_ver = int(fichero_cfg.get('Version_Min_Revit',2024))
add('Fichero','Version Revit','OK' if ver>=min_ver else 'ERROR',f'Revit {ver} (minimo requerido: {min_ver})')

# B: MODELO
is_ws = doc.IsWorkshared
add('Modelo','Workshared','OK' if is_ws else 'ERROR','Modelo compartido activo' if is_ws else 'El modelo NO tiene worksharing activo — activarlo en Colaborar > Gestionar colaboracion')
if is_ws:
    wsets = list(FilteredWorksetCollector(doc).OfKind(WorksetKind.UserWorkset))
    add('Modelo','Worksets','OK',f'{len(wsets)} worksets de usuario: {", ".join(w.Name for w in wsets)}')
warns = list(doc.GetWarnings()); max_w = int(modelo_cfg.get('Max_Advertencias',100))
add('Modelo','Advertencias','OK' if len(warns)<=max_w else 'ERROR',
    f'{len(warns)} advertencias activas (maximo permitido: {max_w}) — revisar y resolver en Gestionar > Advertencias' if len(warns)>max_w else f'{len(warns)} advertencias (dentro del limite de {max_w})')
sp_file = app.SharedParametersFilename or ''
need_sp = str(modelo_cfg.get('Parametros_Compartidos','True'))=='True'
if need_sp:
    if not sp_file.strip(): add('Modelo','Param compartidos','ERROR','No hay fichero de parametros compartidos configurado — ir a Gestionar > Parametros compartidos y seleccionar el fichero .txt del proyecto')
    elif not Path(sp_file).exists(): add('Modelo','Param compartidos','ERROR',f'Ruta configurada pero el fichero NO existe en disco. Ruta: {sp_file}')
    else: add('Modelo','Param compartidos','OK',sp_file)
else: add('Modelo','Param compartidos','OK','No requerido segun configuracion')
all_fams = list(FilteredElementCollector(doc).OfClass(Family).ToElements())
used_ids = set(eid_val(i.GetTypeId()) for i in FilteredElementCollector(doc).OfClass(FamilyInstance).ToElements())
unused_count = sum(1 for f in all_fams if not set(eid_val(s) for s in f.GetFamilySymbolIds()).intersection(used_ids))
add('Modelo','Familias para purgar','OK' if unused_count==0 else 'WARNING',
    f'{unused_count} familias cargadas sin ninguna instancia en el modelo — ejecutar Gestionar > Purgar no usado' if unused_count>0 else 'Todas las familias tienen al menos una instancia')

# C: COORDENADAS
tol = float(coord_cfg.get('Tolerancia_PBP_Survey_mm', 10))
try:
    pbp_list = list(FilteredElementCollector(doc).OfCategory(BuiltInCategory.OST_ProjectBasePoint).ToElements())
    pbp = pbp_list[0] if pbp_list else None
    if pbp:
        ew = ft_to_mm(pbp.get_Parameter(BuiltInParameter.BASEPOINT_EASTWEST_PARAM).AsDouble())
        ns = ft_to_mm(pbp.get_Parameter(BuiltInParameter.BASEPOINT_NORTHSOUTH_PARAM).AsDouble())
        el = ft_to_mm(pbp.get_Parameter(BuiltInParameter.BASEPOINT_ELEVATION_PARAM).AsDouble())
        add('Coordenadas','PBP coords compartidas','OK',f'E/O:{ew:.1f}mm  N/S:{ns:.1f}mm  Elev:{el:.1f}mm')

        for key, val_act in [('E/O', ew), ('N/S', ns), ('Elevacion', el)]:
            exp_raw = coord_cfg.get(key)
            if exp_raw is not None:
                exp_mm = float(exp_raw) / 10.0
                diff = abs(val_act - exp_mm)
                estado = 'OK' if diff <= max(tol, 1) else 'ERROR'
                add('Coordenadas', f'PBP {key} vs BEP',
                    estado,
                    f'Modelo: {val_act:.1f}mm  BEP: {exp_mm:.1f}mm  Diferencia: {diff:.1f}mm'
                    + ('' if estado == 'OK' else ' — actualizar el Punto Base del Proyecto en Revit'))

        ang_p = pbp.get_Parameter(BuiltInParameter.BASEPOINT_ANGLETON_PARAM)
        if ang_p:
            ang_deg = math.degrees(ang_p.AsDouble())
            exp_ang = coord_cfg.get('Angulo Norte real')
            if exp_ang is not None:
                diff_a = abs(ang_deg - float(exp_ang))
                add('Coordenadas', 'Angulo Norte',
                    'OK' if diff_a < 0.01 else 'ERROR',
                    f'Modelo: {ang_deg:.4f}°  BEP: {float(exp_ang):.4f}°  Diferencia: {diff_a:.4f}°')

        try:
            pbp_bp = BasePoint.GetProjectBasePoint(doc)
            pos = pbp_bp.Position
            x_mm = ft_to_mm(pos.X); y_mm = ft_to_mm(pos.Y); z_mm = ft_to_mm(pos.Z)
            dist = (x_mm**2 + y_mm**2 + z_mm**2)**0.5
            st = 'OK' if dist < 1 else 'WARNING'
            add('Coordenadas', 'PBP coincide con Origen Interno', st,
                f'Distancia al origen (0,0,0): {dist:.1f}mm'
                + ('' if dist < 1 else ' — el Punto Base se ha movido del origen interno. Verificar si es intencionado.'))
        except Exception as e2:
            add('Coordenadas', 'PBP vs Origen Interno', 'WARNING', f'No se pudo comprobar: {e2}')
    else:
        add('Coordenadas', 'Project Base Point', 'ERROR', 'Elemento no encontrado en el modelo')
except Exception as e:
    add('Coordenadas', 'Error', 'ERROR', str(e))

try:
    links = list(FilteredElementCollector(doc).OfClass(RevitLinkInstance).ToElements())
    if not links:
        add('Coordenadas', 'Vinculos', 'OK', 'Sin vinculos RVT en el modelo')
    else:
        need_pin = str(coord_cfg.get('Links_Por_Coordenadas', 'True')) == 'True'
        for lnk in links:
            t = lnk.GetTotalTransform(); o = t.Origin; pinned = lnk.Pinned
            st = 'OK' if (not need_pin or pinned) else 'WARNING'
            add('Coordenadas', f'Vinculo: {lnk.Name}', st,
                f'Anclado: {"Si" if pinned else "No — anclarlo una vez verificada la posicion"}  Origen:({ft_to_mm(o.X):.0f},{ft_to_mm(o.Y):.0f},{ft_to_mm(o.Z):.0f})mm')
except Exception as e:
    add('Coordenadas', 'Vinculos', 'ERROR', str(e))

# D: PROYECTO
PI_NAME_MAP={'Project Name':'Nombre de proyecto','Project Number':'Número de proyecto',
    'Client Name':'Nombre de cliente','Author':'Autor','Project Address':'Dirección de proyecto',
    'Project Status':'Estado de proyecto','Organization Name':'Nombre de organización',
    'Building Name':'Nombre del edificio'}
pi=doc.ProjectInformation
for r in proyecto_cfg:
    pname,oblig=r[0],r[1]; val_esp=r[3] if len(r)>3 else None
    p=pi.LookupParameter(PI_NAME_MAP.get(pname,pname)) or pi.LookupParameter(pname)
    if p is None: add('Proyecto',pname,'WARNING',f'Parametro "{pname}" no encontrado en Informacion del proyecto'); continue
    val=get_pval(p)
    if not val.strip(): add('Proyecto',pname,'ERROR' if str(oblig)=='True' else 'OK','Campo obligatorio vacio — rellenar en Gestionar > Informacion del proyecto')
    elif val_esp and str(val).strip()!=str(val_esp).strip(): add('Proyecto',pname,'WARNING',f'Valor actual: "{val}" — esperado: "{val_esp}"')
    else: add('Proyecto',pname,'OK',val)

# E: PLANOS
sheets=list(FilteredElementCollector(doc).OfClass(ViewSheet).ToElements())
if not sheets: add('Planos','Resumen','WARNING','No hay planos en el modelo')
else:
    nom_ok_count=0; nom_err_count=0
    for sheet in sheets:
        snum=sheet.SheetNumber or ''; sname=sheet.Name or ''
        if not re.match(r'^[A-Z]{2,4}-\d{3}$',snum):
            add('Planos','[Nomenclatura]','ERROR',
                f'"{snum} - {sname}": no sigue el formato DISC-NNN (ej: ARQ-001).',eid_val(sheet.Id))
            nom_err_count+=1
        else:
            nom_ok_count+=1
        for r in planos_cfg:
            pname,oblig=r[0],r[1]; val_esp=r[3] if len(r)>3 else None
            if str(oblig)!='True': continue
            val=get_sheet_pval(sheet,pname)
            if val is None: continue
            if not val.strip(): add('Planos',f'[Parametro] {pname}','ERROR',f'"{snum} — {sname}": campo "{pname}" vacio.',eid_val(sheet.Id))
            elif val_esp and str(val).strip()!=str(val_esp).strip():
                add('Planos',f'[Parametro] {pname}','ERROR' if str(oblig)=='True' else 'WARNING',f'"{snum}": valor "{val}" (esperado: "{val_esp}")',eid_val(sheet.Id))
    nom_st='OK' if nom_err_count==0 else ('WARNING' if nom_err_count<len(sheets) else 'ERROR')
    add('Planos','[Nomenclatura] Resumen',nom_st,
        f'{nom_ok_count} planos con nomenclatura correcta — {nom_err_count} con errores')
    n_err=len([x for x in results.get('Planos',[]) if x['Estado']=='ERROR'])
    res_st='OK' if n_err==0 else ('WARNING' if n_err<len(sheets) else 'ERROR')
    add('Planos','[Resumen]',res_st,f'{len(sheets)} planos analizados — {n_err} errores encontrados')

# F: PARAMETROS
print('Parametros...')
for bic_str,param_map in matrix.items():
    if not param_map: continue
    grupo,cat_name=cat_info.get(bic_str,('',''))
    bic=BIC_MAP.get(bic_str)
    if bic is None: continue
    try:
        t_params={k for k,v in param_map.items() if v=='T'}
        e_params={k for k,v in param_map.items() if v=='E'}
        cat_err=0
        if t_params:
            types=list(FilteredElementCollector(doc).OfCategory(bic).WhereElementIsElementType().ToElements())
            if not types:
                for pn in t_params: add('Parametros',f'{cat_name}/{pn}','WARNING',f'Sin tipos de {cat_name} en el modelo')
            else:
                for pn in t_params:
                    no_param=0; empty=0
                    for et in types:
                        p=et.LookupParameter(pn)
                        if p is None: no_param+=1
                        elif not get_pval(p).strip(): empty+=1; add('Parametros',f'{cat_name}/{pn}','ERROR',f'Tipo "{et.Name}": vacio — rellenar en propiedades del tipo',eid_val(et.Id))
                    if no_param>0:
                        add('Parametros',f'{cat_name}/{pn}','ERROR',f'No cargado en {no_param}/{len(types)} tipos — verificar fichero de parametros compartidos y que el parametro esta asignado a esta categoria')
                        cat_err+=no_param
                    cat_err+=empty
        if e_params:
            insts=list(FilteredElementCollector(doc).OfCategory(bic).WhereElementIsNotElementType().ToElements())
            if not insts:
                for pn in e_params: add('Parametros',f'{cat_name}/{pn}','WARNING',f'Sin instancias de {cat_name} en el modelo')
            else:
                for pn in e_params:
                    no_param=0; empty=0
                    for inst in insts:
                        p=inst.LookupParameter(pn)
                        if p is None: no_param+=1
                        elif not get_pval(p).strip(): empty+=1; add('Parametros',f'{cat_name}/{pn}','ERROR',f'Elemento ID {eid_val(inst.Id)}: vacio',eid_val(inst.Id))
                    if no_param>0:
                        add('Parametros',f'{cat_name}/{pn}','ERROR',f'No cargado en {no_param}/{len(insts)} instancias — verificar fichero de parametros compartidos')
                        cat_err+=no_param
                    cat_err+=empty
        n=len(list(FilteredElementCollector(doc).OfCategory(bic).WhereElementIsNotElementType().ToElements()))
        add('Parametros',f'{cat_name}/Resumen','OK' if cat_err==0 else 'ERROR',f'{n} instancias analizadas — {cat_err} problemas')
    except Exception as e: add('Parametros',cat_name,'ERROR',str(e))

# G: NOMENCLATURA TIPOS
print('Nomenclatura tipos...')
for bic_str in matrix:
    grupo,cat_name=cat_info.get(bic_str,('',''))
    disc_code=disciplinas.get(grupo,'')
    valid_codes=[c[1] for c in clasificaciones.get(bic_str,[])]
    bic=BIC_MAP.get(bic_str)
    if bic is None: continue
    try:
        types=list(FilteredElementCollector(doc).OfCategory(bic).WhereElementIsElementType().ToElements())
        n_ok=n_err=0
        for et in types:
            clas_param_val = None
            cp = et.LookupParameter('Editeca_3D_Clasificacion')
            if cp: clas_param_val = get_pval(cp).strip()
            errs = nom_errors(et.Name, disc_code, cat_name, valid_codes, clas_param_val)
            if errs: add('Nomenclatura Tipos',cat_name,'ERROR', f'"{et.Name}": {" | ".join(errs)}',eid_val(et.Id)); n_err+=1
            else: n_ok+=1
        if types: add('Nomenclatura Tipos',f'{cat_name}/Resumen','OK' if n_err==0 else 'ERROR',f'{n_ok} tipos correctos — {n_err} tipos con errores de nomenclatura')
    except Exception as e: add('Nomenclatura Tipos',cat_name,'ERROR',str(e))

# H: NOMENCLATURA FAMILIAS
print('Nomenclatura familias...')
for bic_str in matrix:
    grupo,cat_name=cat_info.get(bic_str,('',''))
    disc_code=disciplinas.get(grupo,'')
    valid_codes=[c[1] for c in clasificaciones.get(bic_str,[])]
    bic=BIC_MAP.get(bic_str)
    if bic is None: continue
    try:
        bic_int=int(bic)
        fams=[f for f in all_fams if f.FamilyCategory is not None and eid_val(f.FamilyCategory.Id)==bic_int]
        n_ok=n_err=0
        for fam in fams:
            errs = nom_errors(fam.Name, disc_code, cat_name, valid_codes)
            if errs: add('Nomenclatura Familias',cat_name,'ERROR', f'"{fam.Name}": {" | ".join(errs)}',eid_val(fam.Id)); n_err+=1
            else: n_ok+=1
        if fams: add('Nomenclatura Familias',f'{cat_name}/Resumen','OK' if n_err==0 else 'ERROR',f'{n_ok} familias correctas — {n_err} con errores')
    except Exception as e: add('Nomenclatura Familias',cat_name,'ERROR',str(e))

# EXCEL
print('Excel...')
S_FILL={'OK':'C6EFCE','ERROR':'FFC7CE','WARNING':'FFEB9C'}; S_FONT={'OK':'006100','ERROR':'9C0006','WARNING':'9C5700'}
out_wb=openpyxl.Workbook(); out_wb.remove(out_wb.active)
for sec,rr in results.items():
    ws=out_wb.create_sheet(sec[:31])
    ws.append(['Check','Estado','Detalle','ElementId'])
    for cell in ws[1]: cell.font=XFont(bold=True,color='FFFFFF'); cell.fill=PatternFill('solid',fgColor='1F3864')
    for row in rr:
        ws.append([row['Check'],row['Estado'],row['Detalle'],row['ElementId']])
        for cell in ws[ws.max_row]:
            cell.fill=PatternFill('solid',fgColor=S_FILL.get(row['Estado'],'FFFFFF'))
            cell.font=XFont(color=S_FONT.get(row['Estado'],'000000'))
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width=min(max(len(str(c.value or '')) for c in col)+2,60)
excel_out=str(Path(OUTPUT_DIR)/f'QC_ModelAudit_{TODAY}.xlsx')
out_wb.save(excel_out)

# HTML
print('HTML...')
def he(s): return str(s).replace('&','&amp;').replace('<','&lt;').replace('>','&gt;')
def badge(st): return f'<span class="b-{st.lower()}">{st}</span>'

GROUPED_SECS=('Parametros','Nomenclatura Tipos','Nomenclatura Familias')
summary={s:{'OK':0,'ERROR':0,'WARNING':0,'total':0} for s in results}
for s,rr in results.items():
    for r in rr:
        if s in GROUPED_SECS and '/Resumen' not in r['Check']: continue
        if s=='Planos' and 'Resumen' not in r['Check']: continue
        summary[s][r['Estado']]+=1; summary[s]['total']+=1

total_ok=sum(v['OK'] for v in summary.values()); total_err=sum(v['ERROR'] for v in summary.values())
total_warn=sum(v['WARNING'] for v in summary.values()); total_all=total_ok+total_err+total_warn

WEIGHTS={'Fichero':1,'Modelo':2,'Coordenadas':2,'Proyecto':1,'Planos':2,'Parametros':3,'Nomenclatura Tipos':2,'Nomenclatura Familias':2}
sw=0; mw=0
for s,c in summary.items():
    w=WEIGHTS.get(s,1); d=c['OK']+c['ERROR']+c['WARNING']
    if d>0: sw+=(c['OK']/d)*w; mw+=w
score=round(sw/mw*10,1) if mw>0 else 0
grade='A' if score>=9 else 'B' if score>=7 else 'C' if score>=5 else 'D' if score>=3 else 'F'
grade_color='#27ae60' if score>=7 else '#e67e22' if score>=5 else '#e74c3c'
model_name=doc.Title.replace('.rvt','')

def sec_cls(c):
    if c['ERROR']>0: return 'err','#e74c3c'
    if c['WARNING']>0: return 'warn','#e67e22'
    return 'ok','#27ae60'

C=62.83; pok=round(total_ok/total_all*100) if total_all else 0; perr=round(total_err/total_all*100) if total_all else 0; pwarn=100-pok-perr
ok_d=round(pok/100*C,1); err_d=round(perr/100*C,1); warn_d=round(pwarn/100*C,1)
err_off=round(15.7-ok_d,1); warn_off=round(15.7-ok_d-err_d,1)
donut=f'<svg width="200" height="200" viewBox="0 0 44 44"><circle cx="22" cy="22" r="16" fill="none" stroke="#ecf0f1" stroke-width="5"/><circle cx="22" cy="22" r="16" fill="none" stroke="#27ae60" stroke-width="5" stroke-dasharray="{ok_d} {round(C-ok_d,1)}" stroke-dashoffset="15.7"/><circle cx="22" cy="22" r="16" fill="none" stroke="#e74c3c" stroke-width="5" stroke-dasharray="{err_d} {round(C-err_d,1)}" stroke-dashoffset="{err_off}"/><circle cx="22" cy="22" r="16" fill="none" stroke="#e67e22" stroke-width="5" stroke-dasharray="{warn_d} {round(C-warn_d,1)}" stroke-dashoffset="{warn_off}"/><text x="22" y="20" text-anchor="middle" font-size="7" font-weight="800" fill="#2c3e50">{pok}%</text><text x="22" y="26" text-anchor="middle" font-size="3.5" fill="#7f8c8d">CORRECTOS</text></svg>'

bars=''.join(f'<div class="br"><div class="bl">{he(s)}</div><div class="bt"><div style="width:{round(c["OK"]/c["total"]*100) if c["total"] else 0}%;background:#27ae60;height:100%"></div><div style="width:{round(c["ERROR"]/c["total"]*100) if c["total"] else 0}%;background:#e74c3c;height:100%"></div><div style="width:{max(0,100-round(c["OK"]/c["total"]*100)-round(c["ERROR"]/c["total"]*100)) if c["total"] else 0}%;background:#e67e22;height:100%"></div></div><div class="bp">{round(c["OK"]/c["total"]*100) if c["total"] else 0}%</div></div>' for s,c in summary.items() if c['total'])

def get_cats(sec):
    return sorted(set(r['Check'].split('/')[0] for r in results.get(sec,[]) if r['Check']))

nav=''
for s,c in summary.items():
    cls,col=sec_cls(c); sid=he(s).replace(' ','_')
    nav+=f'<a href="#{sid}" class="na {cls}">&#x25CF; {he(s)}</a>'
    if s in ('Parametros','Nomenclatura Tipos','Nomenclatura Familias'):
        for cat in get_cats(s):
            cid=f'{sid}_{he(cat).replace(" ","_")}'
            nav+=f'<a href="#{cid}" class="na-sub">&#x25E6; {he(cat)}</a>'

cards='<div class="sg">'
for s,c in summary.items():
    cls,col=sec_cls(c); d=c['total']; op=round(c['OK']/d*100) if d else 0; sid=he(s).replace(' ','_')
    cards+=f'<div class="sc {cls}" onclick="document.getElementById(\'{sid}\').scrollIntoView({{behavior:\'smooth\'}})" style="cursor:pointer"><div class="st">{he(s)}</div><div class="sn2"><span class="b-ok">OK {c["OK"]}</span><span class="b-err">ERR {c["ERROR"]}</span><span class="b-warn">WRN {c["WARNING"]}</span></div><div class="mb"><div style="width:{op}%;background:#27ae60"></div><div style="width:{100-op}%;background:#e0e0e0"></div></div><div class="sp" style="color:{col}">{op}%</div></div>'
cards+='</div>'

def simple_tbl(rr):
    rows=''.join(f'<tr><td>{he(r["Check"])}</td><td>{badge(r["Estado"])}</td><td>{he(r["Detalle"])}</td><td><span class="eid">{he(r["ElementId"])}</span></td></tr>' for r in rr)
    return f'<table class="dt"><thead><tr><th>Check</th><th>Estado</th><th>Detalle</th><th>ElementId</th></tr></thead><tbody>{rows}</tbody></table>'

def planos_tbl(rr):
    nom_det=[r for r in rr if r['Check']=='[Nomenclatura]']
    nom_res=next((r for r in rr if r['Check']=='[Nomenclatura] Resumen'),None)
    par_det=[r for r in rr if r['Check'].startswith('[Parametro]')]
    overall=next((r for r in rr if r['Check']=='[Resumen]'),None)
    def subgroup(label, det, res_row):
        st=res_row['Estado'] if res_row else ('ERROR' if any(r['Estado']=='ERROR' for r in det) else ('WARNING' if det else 'OK'))
        _,col=sec_cls({'ERROR':int(st=='ERROR'),'WARNING':int(st=='WARNING'),'OK':int(st=='OK')})
        sum_txt=res_row['Detalle'] if res_row else (f'{len(det)} errores' if det else 'Sin errores')
        _nums=re.findall(r'\d+',sum_txt)
        n_ok=int(_nums[0]) if _nums and st=='OK' else 0
        n_err=sum(1 for r in det if r['Estado']=='ERROR'); n_warn=sum(1 for r in det if r['Estado']=='WARNING')
        if not det:
            return f'<div class="cb"><div class="ch" style="border-left:3px solid {col};cursor:default"><span class="cn">{he(label)}</span><span class="cs">{he(sum_txt)}</span><span style="display:flex;gap:4px"><span class="b-ok">OK {n_ok}</span><span class="b-err">ERR {n_err}</span><span class="b-warn">WRN {n_warn}</span></span></div></div>'
        det_rows=''.join(f'<tr><td>{he(r["Check"].replace("[Parametro] ",""))}</td><td>{badge(r["Estado"])}</td><td>{he(r["Detalle"])}</td><td><span class="eid">{he(r["ElementId"])}</span></td></tr>' for r in det)
        return f'<div class="cb"><div class="ch" style="border-left:3px solid {col}" onclick="var b=this.nextElementSibling;b.style.display=b.style.display==\'none\'?\'block\':\'none\'"><span class="cn">{he(label)}</span><span class="cs">{he(sum_txt)}</span><span style="display:flex;gap:4px"><span class="b-ok">OK {n_ok}</span><span class="b-err">ERR {n_err}</span><span class="b-warn">WRN {n_warn}</span></span></div><div class="cbody" style="display:none"><table class="dt"><thead><tr><th>Check</th><th>Estado</th><th>Detalle</th><th>ElementId</th></tr></thead><tbody>{det_rows}</tbody></table></div></div>'
    return subgroup('Nomenclatura', nom_det, nom_res) + subgroup('Parametros', par_det, overall)

def grouped_tbl(rr, sec_id):
    cats={}
    for r in rr: cats.setdefault(r['Check'].split('/')[0],[]).append(r)
    html=''
    for cat,rows in cats.items():
        cid=f'{sec_id}_{he(cat).replace(" ","_")}'
        res=[r for r in rows if '/Resumen' in r['Check']]
        det=[r for r in rows if '/Resumen' not in r['Check']]
        st=res[0]['Estado'] if res else ('ERROR' if any(r['Estado']=='ERROR' for r in det) else ('WARNING' if any(r['Estado']=='WARNING' for r in det) else 'OK'))
        _,col=sec_cls({'ERROR':int(st=='ERROR'),'WARNING':int(st=='WARNING'),'OK':int(st=='OK')})
        sum_txt=res[0]['Detalle'] if res else ''
        if not det and res:
            _nums=re.findall(r'\d+',res[0]['Detalle'])
            n_ok=int(_nums[0]) if _nums and res[0]['Estado']=='OK' else 0
            n_err=int(_nums[0]) if _nums and res[0]['Estado']=='ERROR' else 0
            n_warn=0
        else:
            n_err=sum(1 for r in det if r['Estado']=='ERROR'); n_warn=sum(1 for r in det if r['Estado']=='WARNING')
            if res:
                _nums=re.findall(r'\d+',res[0]['Detalle'])
                n_ok=int(_nums[0]) if _nums else sum(1 for r in det if r['Estado']=='OK')
            else:
                n_ok=sum(1 for r in det if r['Estado']=='OK')
        if not det:
            html+=f'<div class="cb" id="{cid}"><div class="ch" style="border-left:3px solid {col};cursor:default"><span class="cn">{he(cat)}</span><span class="cs">{he(sum_txt)}</span><span style="display:flex;gap:4px"><span class="b-ok">OK {n_ok}</span><span class="b-err">ERR {n_err}</span><span class="b-warn">WRN {n_warn}</span></span></div></div>'
        else:
            det_rows=''.join(f'<tr><td class="chk-col">{he(r["Check"].split("/")[-1])}</td><td>{badge(r["Estado"])}</td><td>{he(r["Detalle"])}</td><td><span class="eid">{he(r["ElementId"])}</span></td></tr>' for r in det)
            html+=f'<div class="cb" id="{cid}"><div class="ch" style="border-left:3px solid {col}" onclick="var b=this.nextElementSibling;b.style.display=b.style.display==\'none\'?\'block\':\'none\'"><span class="cn">{he(cat)}</span><span class="cs">{he(sum_txt)}</span><span style="display:flex;gap:4px"><span class="b-ok">OK {n_ok}</span><span class="b-err">ERR {n_err}</span><span class="b-warn">WRN {n_warn}</span></span></div><div class="cbody" style="display:none"><table class="dt"><thead><tr><th>Parametro</th><th>Estado</th><th>Detalle</th><th>ElementId</th></tr></thead><tbody>{det_rows}</tbody></table></div></div>'
    return html

detail=''
for s,rr in results.items():
    cls,col=sec_cls(summary[s]); sid=he(s).replace(' ','_')
    if s=='Planos': tbl=planos_tbl(rr)
    elif s in ('Parametros','Nomenclatura Tipos','Nomenclatura Familias'): tbl=grouped_tbl(rr,sid)
    else: tbl=simple_tbl(rr)
    detail+=f'<div class="ss" id="{sid}"><div class="sh" style="border-left:4px solid {col}"><div><h2>{he(s)}</h2><span class="mc">{summary[s]["total"]} checks</span></div><div class="ck"><span class="b-ok">OK {summary[s]["OK"]}</span><span class="b-err">ERR {summary[s]["ERROR"]}</span><span class="b-warn">WRN {summary[s]["WARNING"]}</span></div></div>{tbl}</div>'

css=':root{--bg:#f4f6f9;--card:#fff;--hdr:#1a2332;--text:#2c3e50;--muted:#7f8c8d;--bdr:#dde3ec;--ok:#27ae60;--err:#e74c3c;--warn:#e67e22}*{box-sizing:border-box;margin:0;padding:0}body{font-family:"Segoe UI",Arial,sans-serif;background:var(--bg);color:var(--text);font-size:14px}header{background:var(--hdr);color:#fff;padding:16px 28px;display:flex;align-items:center;justify-content:space-between;position:sticky;top:0;z-index:100}header h1{font-size:16px;font-weight:600}header .m{font-size:11px;color:#9ab;text-align:right;line-height:1.8}.lay{display:flex}nav{width:200px;flex-shrink:0;background:#1e293b;min-height:calc(100vh - 52px);position:sticky;top:52px;height:calc(100vh - 52px);overflow-y:auto}.sb{text-align:center;padding:18px 8px 14px;border-bottom:1px solid #2d3f55}.sn{font-size:42px;font-weight:800;line-height:1}.sg2{font-size:12px;font-weight:600;margin-top:2px}.ss2{font-size:10px;color:#9ab;margin-top:3px}.nt{font-size:10px;text-transform:uppercase;letter-spacing:.8px;color:#9ab;padding:12px 14px 5px}.na{display:block;padding:6px 14px;font-size:12px;text-decoration:none;color:#c8d4e3;transition:background .15s}.na:hover{background:#2d3f55}.na.ok{color:#6fcf97}.na.err{color:#f87171}.na.warn{color:#fbbf24}.na-sub{display:block;padding:4px 14px 4px 28px;font-size:11px;text-decoration:none;color:#7a8fa6;transition:background .15s}.na-sub:hover{background:#2d3f55;color:#c8d4e3}main{flex:1;max-width:1080px;padding:22px;overflow:auto}.kr{display:grid;grid-template-columns:repeat(4,1fr);gap:11px;margin-bottom:18px}.kp{background:var(--card);border-radius:9px;padding:14px 18px;border-left:4px solid var(--muted);box-shadow:0 1px 4px rgba(0,0,0,.07)}.kp.e{border-color:var(--err)}.kp.w{border-color:var(--warn)}.kp.o{border-color:var(--ok)}.kp .v{font-size:32px;font-weight:700;line-height:1;margin-bottom:3px}.kp .l{font-size:11px;color:var(--muted);text-transform:uppercase;letter-spacing:.4px}.cr{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:18px}.cd{background:var(--card);border-radius:9px;padding:18px;box-shadow:0 1px 4px rgba(0,0,0,.07)}.cd h3{font-size:11px;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.5px;margin-bottom:12px}.dw{display:flex;align-items:center;gap:16px;justify-content:center}.lr{display:flex;flex-direction:column;gap:6px}.li{display:flex;align-items:center;gap:6px;font-size:12px}.ld{width:9px;height:9px;border-radius:50%;flex-shrink:0}.br{display:flex;align-items:center;gap:7px;margin-bottom:6px}.bl{width:130px;font-size:11px;text-align:right;flex-shrink:0}.bt{flex:1;display:flex;height:16px;border-radius:3px;overflow:hidden;background:#ecf0f1}.bp{width:30px;font-size:11px;font-weight:600;color:var(--muted)}.sg{display:grid;grid-template-columns:repeat(auto-fill,minmax(175px,1fr));gap:9px;margin-bottom:18px}.sc{background:var(--card);border-radius:8px;padding:11px;box-shadow:0 1px 3px rgba(0,0,0,.07);border-top:3px solid #ccc;transition:box-shadow .15s}.sc:hover{box-shadow:0 3px 10px rgba(0,0,0,.12)}.sc.ok{border-color:var(--ok)}.sc.err{border-color:var(--err)}.sc.warn{border-color:var(--warn)}.st{font-size:11px;font-weight:600;margin-bottom:5px}.sn2{display:flex;gap:3px;flex-wrap:wrap;margin-bottom:5px}.mb{display:flex;height:4px;border-radius:2px;overflow:hidden;margin-bottom:3px;background:#eee}.sp{font-size:12px;font-weight:700}.ss{margin-bottom:20px}.sh{display:flex;align-items:center;justify-content:space-between;padding:11px 15px;background:var(--card);border-radius:9px 9px 0 0;border-bottom:1px solid var(--bdr)}.sh h2{font-size:12px;font-weight:700;text-transform:uppercase;letter-spacing:.4px}.mc{font-size:11px;color:var(--muted);margin-left:6px}.ck{display:flex;gap:5px}.sub-h{padding:9px 13px 5px;font-size:11px;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.4px;background:#f8fafc;border-bottom:1px solid var(--bdr)}.cb{border:1px solid var(--bdr);border-radius:5px;margin:6px;overflow:hidden}.ch{display:flex;align-items:center;padding:8px 11px;background:#f8fafc;cursor:pointer;gap:8px;user-select:none}.ch:hover{background:#f0f3f7}.cn{font-weight:600;font-size:12px;flex:1;min-width:80px}.cs{font-size:11px;color:var(--muted);flex:2}.cbody{}.chk-col{font-family:monospace;font-size:11px;color:#2c3e50;white-space:nowrap;font-weight:600}.b-ok{display:inline-block;padding:2px 6px;border-radius:20px;font-size:10px;font-weight:600;background:#eafaf1;color:var(--ok);border:1px solid #a9dfbf;white-space:nowrap}.b-err,.b-error{display:inline-block;padding:2px 6px;border-radius:20px;font-size:10px;font-weight:600;background:#fdf2f2;color:var(--err);border:1px solid #f5c6cb;white-space:nowrap}.b-warn,.b-warning{display:inline-block;padding:2px 6px;border-radius:20px;font-size:10px;font-weight:600;background:#fef9e7;color:var(--warn);border:1px solid #fad7a0;white-space:nowrap}table.dt{border-collapse:collapse;width:100%;font-size:12px}table.dt th{background:#f0f3f7;padding:6px 10px;text-align:left;font-size:10px;text-transform:uppercase;letter-spacing:.3px;color:var(--muted);border-bottom:2px solid var(--bdr)}table.dt td{padding:5px 10px;border-bottom:1px solid var(--bdr);vertical-align:middle}table.dt tr:last-child td{border-bottom:none}table.dt tr:hover td{background:#f8fafc}.eid{font-family:monospace;font-size:10px;color:var(--muted)}footer{text-align:center;padding:14px;font-size:11px;color:#aaa;border-top:1px solid var(--bdr);margin-top:16px}'

html=f'<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8"><title>QC — {he(model_name)}</title><style>{css}</style></head><body><header><div><h1>QC Model Audit — {he(model_name)}</h1><div style="font-size:11px;color:#9ab;margin-top:2px">Revit {app.VersionNumber} · {len(results)} secciones</div></div><div class="m"><div>{DATE_STR}</div><div>PyNET · Editeca</div></div></header><div class="lay"><nav><div class="sb"><div class="sn" style="color:{grade_color}">{score}</div><div class="sg2" style="color:{grade_color}">Nota {grade}</div><div class="ss2">Calidad /10</div></div><div class="nt">Indice</div><a href="#resumen" class="na" style="color:#9ab">&#x25A3; Resumen</a>{nav}</nav><main><div id="resumen"><div class="kr"><div class="kp"><div class="v">{total_all}</div><div class="l">Checks</div></div><div class="kp e"><div class="v" style="color:var(--err)">{total_err}</div><div class="l">Errores</div></div><div class="kp w"><div class="v" style="color:var(--warn)">{total_warn}</div><div class="l">Advertencias</div></div><div class="kp o"><div class="v" style="color:var(--ok)">{total_ok}</div><div class="l">Correctos</div></div></div><div class="cr"><div class="cd"><h3>Distribucion</h3><div class="dw">{donut}<div class="lr"><div class="li"><div class="ld" style="background:var(--ok)"></div><b>{total_ok}</b>&nbsp;Correctos ({pok}%)</div><div class="li"><div class="ld" style="background:var(--err)"></div><b>{total_err}</b>&nbsp;Errores ({perr}%)</div><div class="li"><div class="ld" style="background:var(--warn)"></div><b>{total_warn}</b>&nbsp;Advertencias ({pwarn}%)</div></div></div></div><div class="cd"><h3>Cumplimiento por seccion</h3>{bars}</div></div>{cards}</div>{detail}</main></div><footer>PyNET Platform · {he(model_name)} · {DATE_STR} · Nota QC: {score}/10 ({grade})</footer></body></html>'

html_out=str(Path(OUTPUT_DIR)/f'QC_ModelAudit_{TODAY}.html')
with open(html_out,'w',encoding='utf-8') as f: f.write(html)
webbrowser.open(f'file:///{html_out.replace(chr(92), "/")}')
print(f'OK={total_ok} ERR={total_err} WARN={total_warn} Nota={score}/10 ({grade})')
ia_Result={'nota':score,'grade':grade,'OK':total_ok,'ERROR':total_err,'WARNING':total_warn,'html':html_out,'excel':excel_out}
