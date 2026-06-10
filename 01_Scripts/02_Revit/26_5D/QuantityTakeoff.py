
import openpyxl, webbrowser, math
from openpyxl.styles import Font as XFont, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from System import Environment
from datetime import datetime
from collections import OrderedDict

import clr
clr.AddReference("RevitAPI")
from Autodesk.Revit.DB import *

doc = __revit__.ActiveUIDocument.Document  # type: ignore
FT2_TO_M2 = 0.0929
FT_TO_M   = 0.3048
AREA_CATS   = {"Suelos", "Techos", "Cubiertas"}
LENGTH_CATS = {"Muros"}

BIC_MAP = {
    "Muros":               BuiltInCategory.OST_Walls,
    "Suelos":              BuiltInCategory.OST_Floors,
    "Techos":              BuiltInCategory.OST_Ceilings,
    "Cubiertas":           BuiltInCategory.OST_Roofs,
    "Puertas":             BuiltInCategory.OST_Doors,
    "Ventanas":            BuiltInCategory.OST_Windows,
    "Luminarias":          BuiltInCategory.OST_LightingFixtures,
    "Aparatos sanitarios": BuiltInCategory.OST_PlumbingFixtures,
}

CAP_COLORS = {
    "Cerramientos":             "#2980b9",
    "Forjados y Pavimentos":    "#27ae60",
    "Techos":                   "#c9a227",
    "Carpintería":              "#e67e22",
    "Instalaciones Eléctricas": "#8e44ad",
    "Instalaciones Sanitarias": "#c0392b",
}
CAP_LIGHT = {
    "Cerramientos":             "#d6eaf8",
    "Forjados y Pavimentos":    "#d5f5e3",
    "Techos":                   "#fef9e7",
    "Carpintería":              "#fdebd0",
    "Instalaciones Eléctricas": "#ebf5fb",
    "Instalaciones Sanitarias": "#f5eef8",
}

PARAM_RECURSO = "PyNET_5D_CodigoRecurso"

desktop = Environment.GetFolderPath(Environment.SpecialFolder.Desktop)
ref_path = Path(desktop) / "5D_Recursos_Referencia.xlsx"
TODAY = datetime.now().strftime("%Y%m%d_%H%M%S")
DATE_STR = datetime.now().strftime("%d/%m/%Y %H:%M")
MODEL_NAME = doc.Title.replace(".rvt", "")

# ── 1. Load reference ──────────────────────────────────────────────────────
wb_ref = openpyxl.load_workbook(str(ref_path), data_only=True)
ws_ref = wb_ref.active
resource_lookup = {}  # code -> {capitulo, desc, unit, price}
for row in ws_ref.iter_rows(min_row=2, values_only=True):
    if not row[0]: continue
    vals = [str(v).strip() if v is not None else "" for v in row[:5]]
    code, cap, desc, unit, price = vals
    if code:
        resource_lookup[code] = {"code": code, "capitulo": cap, "desc": desc, "unit": unit, "price": float(price) if price else 0.0}

# ── 2. Collect quantities with element breakdown ───────────────────────────
budget = {}
for cat_name, bic in BIC_MAP.items():
    type_code_map = {}
    for t in FilteredElementCollector(doc).OfCategory(bic).WhereElementIsElementType().ToElements():
        p = t.LookupParameter(PARAM_RECURSO)
        if p and p.HasValue:
            code = (p.AsString() or "").strip()
            res = resource_lookup.get(code)
            if res:
                type_code_map[int(t.Id.Value)] = res
    if not type_code_map: continue
    for el in FilteredElementCollector(doc).OfCategory(bic).WhereElementIsNotElementType().ToElements():
        tid = int(el.GetTypeId().Value)
        res = type_code_map.get(tid)
        if not res: continue
        code = res["code"]
        if code not in budget:
            budget[code] = {**res, "qty": 0.0, "elements": []}
        if cat_name in LENGTH_CATS:
            p = el.get_Parameter(BuiltInParameter.CURVE_ELEM_LENGTH)
            qty = round(p.AsDouble() * FT_TO_M, 2) if p and p.HasValue else 0.0
        elif cat_name in AREA_CATS:
            p = el.get_Parameter(BuiltInParameter.HOST_AREA_COMPUTED)
            qty = round(p.AsDouble() * FT2_TO_M2, 2) if p and p.HasValue else 0.0
        else:
            qty = 1.0
        budget[code]["qty"] += qty
        budget[code]["elements"].append({"eid": int(el.Id.Value), "qty": qty})

for v in budget.values():
    v["elements"].sort(key=lambda x: x["qty"], reverse=True)
    v["qty"] = round(v["qty"], 2)

chapters = OrderedDict()
for r in sorted(budget.values(), key=lambda x: (x["capitulo"], x["code"])):
    chapters.setdefault(r["capitulo"], []).append(r)

cap_totals = OrderedDict((cap, round(sum(r["qty"] * r["price"] for r in items), 2)) for cap, items in chapters.items())
grand_total = round(sum(cap_totals.values()), 2)
total_elements = sum(len(r["elements"]) for r in budget.values())

print(f"Grand total: {grand_total:,.2f} EUR | {len(budget)} resources | {total_elements} elements")

# ── 3. Excel output ────────────────────────────────────────────────────────
xls_path = Path(desktop) / f"5D_Presupuesto_{TODAY}.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Mediciones"

def brd(top="thin", bot="thin"):
    s = lambda c, w: Side(style=w, color=c)
    return Border(top=s("CCCCCC", top), bottom=s("333333" if bot=="medium" else "CCCCCC", bot),
                  left=s("CCCCCC","thin"), right=s("CCCCCC","thin"))

hdr_fill = PatternFill("solid", fgColor="1F4E79")
hdr_font = XFont(bold=True, color="FFFFFF", size=11)
for col, h in enumerate(["Código BEDEC","Descripción / ElementID","Ud.","Medición","P. Unitario (€)","Total (€)"], 1):
    c = ws.cell(row=1, column=col, value=h)
    c.fill = hdr_fill; c.font = hdr_font
    c.alignment = Alignment(horizontal="center", vertical="center"); c.border = brd()
ws.row_dimensions[1].height = 22
for i, w in enumerate([14,55,6,12,18,14], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"

cap_totals_xls = {}
row_i = 2
for cap, resources in chapters.items():
    cap_hex = CAP_COLORS.get(cap,"333333").lstrip("#")
    cap_fill = PatternFill("solid", fgColor=cap_hex)
    cap_sub = cap_totals[cap]
    cap_totals_xls[cap] = cap_sub
    ws.merge_cells(start_row=row_i, start_column=1, end_row=row_i, end_column=5)
    c = ws.cell(row=row_i, column=1, value=f"  {cap.upper()}")
    c.fill = cap_fill; c.font = XFont(bold=True, color="FFFFFF", size=12); c.alignment = Alignment(vertical="center"); c.border = brd()
    c = ws.cell(row=row_i, column=6, value=round(cap_sub,2))
    c.fill = cap_fill; c.font = XFont(bold=True, color="FFFFFF", size=12); c.number_format = '#,##0.00 "€"'; c.border = brd()
    ws.row_dimensions[row_i].height = 20; row_i += 1

    for r in resources:
        light = CAP_LIGHT.get(cap,"F0F0F0").lstrip("#")
        p_fill = PatternFill("solid", fgColor=light)
        total_p = round(r["qty"] * r["price"], 2)
        for col, v in enumerate([r["code"], r["desc"], r["unit"], r["qty"], r["price"], total_p], 1):
            c = ws.cell(row=row_i, column=col, value=v)
            c.fill = p_fill; c.border = brd(); c.font = XFont(bold=True, size=10)
            c.alignment = Alignment(vertical="center")
            if col in (4,5,6): c.number_format = '#,##0.00'
        ws.row_dimensions[row_i].height = 17; row_i += 1

        for el in r["elements"]:
            el_fill = PatternFill("solid", fgColor="FAFAFA")
            c = ws.cell(row=row_i, column=1, value=el["eid"])
            c.fill = el_fill; c.border = brd(); c.font = XFont(size=9, color="666666"); c.alignment = Alignment(horizontal="right")
            for col in [2,3,5,6]:
                ws.cell(row=row_i, column=col).fill = el_fill; ws.cell(row=row_i, column=col).border = brd()
            c = ws.cell(row=row_i, column=4, value=el["qty"])
            c.fill = el_fill; c.border = brd(); c.number_format = '#,##0.00'; c.font = XFont(size=9, color="444444")
            ws.row_dimensions[row_i].height = 14; row_i += 1
        row_i += 1

last = row_i
for col in range(1,8):
    ws.cell(row=last, column=col).fill = PatternFill("solid", fgColor="1F4E79"); ws.cell(row=last, column=col).border = brd("medium","medium")
ws.cell(row=last, column=3, value="TOTAL PRESUPUESTO").font = XFont(bold=True, color="FFFFFF", size=12)
c = ws.cell(row=last, column=6, value=grand_total); c.font = XFont(bold=True, color="FFFFFF", size=12); c.number_format = '#,##0.00 "€"'

ws2 = wb.create_sheet("Resumen")
for col, h in enumerate(["Capítulo","Total (€)","% s/Total"], 1):
    c = ws2.cell(row=1, column=col, value=h); c.fill = hdr_fill; c.font = hdr_font; c.alignment = Alignment(horizontal="center")
for i, (cap, sub) in enumerate(cap_totals.items(), 2):
    pct = round(sub / grand_total * 100, 1) if grand_total else 0
    fill = PatternFill("solid", fgColor=CAP_LIGHT.get(cap,"FFFFFF").lstrip("#"))
    ws2.cell(row=i, column=1, value=cap).fill = fill; ws2.cell(row=i, column=1).font = XFont(bold=True)
    c = ws2.cell(row=i, column=2, value=round(sub,2)); c.number_format = '#,##0.00 "€"'; c.fill = fill
    c = ws2.cell(row=i, column=3, value=pct); c.number_format = '0.0"%"'; c.fill = fill
lr = len(cap_totals) + 2
ws2.cell(row=lr, column=1, value="TOTAL").font = XFont(bold=True, color="FFFFFF", size=12)
for col in range(1,4): ws2.cell(row=lr,column=col).fill = PatternFill("solid",fgColor="1F4E79")
c = ws2.cell(row=lr,column=2, value=grand_total); c.number_format = '#,##0.00 "€"'; c.font = XFont(bold=True,color="FFFFFF",size=12)
ws2.cell(row=lr,column=3, value=100.0).number_format = '0.0"%"'
for col, w in enumerate([30,18,12], 1): ws2.column_dimensions[get_column_letter(col)].width = w

wb.save(str(xls_path))
print(f"Excel: {xls_path}")

# ── 4. HTML output ─────────────────────────────────────────────────────────
def he(s): return str(s).replace('&','&amp;').replace('<','&lt;').replace('>','&gt;').replace('"','&quot;')
def fmt_eur(v): return f"{v:,.2f} €".replace(",","X").replace(".",",").replace("X",".")

C_D = 100.53  # 2*pi*16
def donut_svg():
    circles = '<circle cx="22" cy="22" r="16" fill="none" stroke="#ecf0f1" stroke-width="5"/>'
    cumulative = 0.0
    for cap, sub in cap_totals.items():
        pct = sub / grand_total if grand_total else 0
        seg = pct * C_D
        offset = C_D / 4 - cumulative
        color = CAP_COLORS.get(cap, "#999")
        circles += f'<circle cx="22" cy="22" r="16" fill="none" stroke="{color}" stroke-width="5" stroke-dasharray="{seg:.2f} {C_D-seg:.2f}" stroke-dashoffset="{offset:.2f}"/>'
        cumulative += seg
    label_top = fmt_eur(grand_total)
    return f'<svg width="210" height="210" viewBox="0 0 44 44">{circles}<text x="22" y="20.5" text-anchor="middle" font-size="3.8" font-weight="800" fill="#2c3e50">TOTAL</text><text x="22" y="25" text-anchor="middle" font-size="3.2" font-weight="600" fill="#555">{he(label_top)}</text></svg>'

legend = "".join(f'<div class="li"><div class="ld" style="background:{CAP_COLORS.get(c,"#999")}"></div><span>{he(c)}</span><b style="margin-left:auto">{fmt_eur(v)}</b></div>' for c,v in cap_totals.items())

max_cap = max(cap_totals.values()) if cap_totals else 1
bars_html = "".join(
    f'<div class="br"><div class="bl">{he(c)}</div><div class="bt"><div style="width:{round(v/max_cap*100)}%;background:{CAP_COLORS.get(c,"#999")};height:100%;border-radius:3px"></div></div><div class="bv">{fmt_eur(v)}</div></div>'
    for c,v in cap_totals.items()
)

cap_cards = "".join(
    f'<div class="sc" style="border-top:3px solid {CAP_COLORS.get(c,"#ccc")}" onclick="document.getElementById(\'cap_{i}\').scrollIntoView({{behavior:\'smooth\'}})" >'
    f'<div class="st">{he(c)}</div>'
    f'<div class="sv">{fmt_eur(v)}</div>'
    f'<div class="sp">{round(v/grand_total*100,1) if grand_total else 0}% del total</div>'
    f'</div>'
    for i,(c,v) in enumerate(cap_totals.items())
)

detail_html = ""
for cap_idx, (cap, resources) in enumerate(chapters.items()):
    cap_color = CAP_COLORS.get(cap, "#333")
    cap_sub = cap_totals[cap]
    n_resources = len(resources)
    n_els = sum(len(r["elements"]) for r in resources)

    partidas_html = ""
    for r in resources:
        total_p = round(r["qty"] * r["price"], 2)
        el_rows = "".join(
            f'<tr><td class="eid-col">{el["eid"]}</td><td></td><td class="num">{el["qty"]:,.2f}</td><td></td><td></td></tr>'
            for el in r["elements"]
        )
        total_row = f'<tr class="total-row"><td></td><td class="total-label">TOTAL PARTIDA</td><td class="num">{r["qty"]:,.2f} {he(r["unit"])}</td><td class="num">{r["price"]:,.2f} €</td><td class="num total-val">{fmt_eur(total_p)}</td></tr>'
        partidas_html += (
            f'<div class="pb">'
            f'<div class="ph" onclick="var b=this.nextElementSibling;b.style.display=b.style.display==\'none\'?\'block\':\'none\'">'
            f'<span class="pc">{he(r["code"])}</span>'
            f'<span class="pd">{he(r["desc"])}</span>'
            f'<span class="pu">{he(r["unit"])}</span>'
            f'<span class="pq">{r["qty"]:,.2f}</span>'
            f'<span class="pp">{r["price"]:,.2f} €</span>'
            f'<span class="pt">{fmt_eur(total_p)}</span>'
            f'</div>'
            f'<div class="pbody" style="display:none">'
            f'<table class="et"><thead><tr><th>ElementID</th><th>Descripción</th><th>Medición</th><th>P. Unit.</th><th>Total</th></tr></thead>'
            f'<tbody>{el_rows}{total_row}</tbody></table>'
            f'</div></div>'
        )

    detail_html += (
        f'<div class="cs" id="cap_{cap_idx}">'
        f'<div class="ch" style="border-left:5px solid {cap_color}">'
        f'<div><h2 style="color:{cap_color}">{he(cap)}</h2>'
        f'<span class="mc">{n_resources} partidas · {n_els} elementos</span></div>'
        f'<div class="cv">{fmt_eur(cap_sub)}</div>'
        f'</div>'
        f'<div class="p-hdr"><span>Código</span><span>Descripción</span><span>Ud.</span><span class="num">Cantidad</span><span class="num">P. Unit.</span><span class="num">Total</span></div>'
        f'{partidas_html}'
        f'</div>'
    )

css = """:root{--bg:#f4f6f9;--card:#fff;--hdr:#1a2332;--text:#2c3e50;--muted:#7f8c8d;--bdr:#dde3ec}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:"Segoe UI",Arial,sans-serif;background:var(--bg);color:var(--text);font-size:14px}
header{background:var(--hdr);color:#fff;padding:16px 28px;display:flex;align-items:center;justify-content:space-between;position:sticky;top:0;z-index:100}
header h1{font-size:17px;font-weight:700}
.lay{display:flex}
nav{width:210px;flex-shrink:0;background:#1e293b;min-height:calc(100vh - 54px);position:sticky;top:54px;height:calc(100vh - 54px);overflow-y:auto}
.sb{text-align:center;padding:18px 8px 14px;border-bottom:1px solid #2d3f55}
.sn{font-size:28px;font-weight:800;color:#f0f0f0;line-height:1.2}
.ss2{font-size:10px;color:#9ab;margin-top:3px}
.nt{font-size:10px;text-transform:uppercase;letter-spacing:.8px;color:#9ab;padding:12px 14px 5px}
.na{display:block;padding:7px 14px;font-size:12px;text-decoration:none;color:#c8d4e3;transition:background .15s;border-left:3px solid transparent}
.na:hover{background:#2d3f55}
main{flex:1;padding:22px;overflow:auto;max-width:1100px}
.kr{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:20px}
.kp{background:var(--card);border-radius:9px;padding:14px 18px;border-left:4px solid #ccc;box-shadow:0 1px 4px rgba(0,0,0,.07)}
.kp .v{font-size:28px;font-weight:800;line-height:1;margin-bottom:3px}
.kp .l{font-size:11px;color:var(--muted);text-transform:uppercase;letter-spacing:.4px}
.cr{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:20px}
.cd{background:var(--card);border-radius:9px;padding:18px;box-shadow:0 1px 4px rgba(0,0,0,.07)}
.cd h3{font-size:11px;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.5px;margin-bottom:14px}
.dw{display:flex;align-items:center;gap:18px}
.lr{display:flex;flex-direction:column;gap:7px;flex:1}
.li{display:flex;align-items:center;gap:8px;font-size:12px}
.ld{width:10px;height:10px;border-radius:50%;flex-shrink:0}
.br{display:flex;align-items:center;gap:8px;margin-bottom:8px}
.bl{width:160px;font-size:11px;text-align:right;flex-shrink:0;color:var(--text)}
.bt{flex:1;height:18px;background:#ecf0f1;border-radius:3px;overflow:hidden}
.bv{width:100px;font-size:11px;font-weight:600;color:var(--muted)}
.sg{display:grid;grid-template-columns:repeat(auto-fill,minmax(170px,1fr));gap:10px;margin-bottom:20px}
.sc{background:var(--card);border-radius:9px;padding:13px;box-shadow:0 1px 3px rgba(0,0,0,.07);cursor:pointer;transition:box-shadow .15s}
.sc:hover{box-shadow:0 4px 12px rgba(0,0,0,.13)}
.st{font-size:11px;font-weight:700;margin-bottom:6px;color:var(--muted);text-transform:uppercase;letter-spacing:.3px}
.sv{font-size:18px;font-weight:800;margin-bottom:3px}
.sp{font-size:11px;color:var(--muted)}
.cs{background:var(--card);border-radius:9px;margin-bottom:20px;box-shadow:0 1px 4px rgba(0,0,0,.07);overflow:hidden}
.ch{display:flex;align-items:center;justify-content:space-between;padding:14px 18px;border-bottom:1px solid var(--bdr)}
.ch h2{font-size:14px;font-weight:700}
.mc{font-size:11px;color:var(--muted);margin-top:3px}
.cv{font-size:22px;font-weight:800}
.p-hdr{display:grid;grid-template-columns:120px 1fr 50px 90px 90px 110px;gap:8px;padding:7px 14px;background:#f0f3f7;font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.3px;color:var(--muted);border-bottom:1px solid var(--bdr)}
.pb{border-bottom:1px solid var(--bdr)}
.ph{display:grid;grid-template-columns:120px 1fr 50px 90px 90px 110px;gap:8px;padding:9px 14px;cursor:pointer;align-items:center;transition:background .12s}
.ph:hover{background:#f8fafc}
.pc{font-family:monospace;font-size:11px;font-weight:700;color:#2c3e50}
.pd{font-size:12px}
.pu{font-size:11px;color:var(--muted)}
.pq,.pp,.pt{font-size:12px;text-align:right}
.pt{font-weight:700}
.pbody{padding:0 14px 10px}
table.et{border-collapse:collapse;width:100%;font-size:11px;margin-top:6px}
table.et th{background:#f8fafc;padding:5px 8px;text-align:left;font-size:10px;text-transform:uppercase;letter-spacing:.3px;color:var(--muted);border-bottom:1px solid var(--bdr)}
table.et td{padding:4px 8px;border-bottom:1px solid #f0f0f0}
table.et .eid-col{font-family:monospace;font-size:10px;color:#888}
table.et .num{text-align:right}
table.et .total-row td{font-weight:700;background:#f4f6f9;border-top:2px solid var(--bdr)}
table.et .total-val{color:#1a2332}
footer{text-align:center;padding:14px;font-size:11px;color:#aaa;border-top:1px solid var(--bdr);margin-top:16px}
"""

nav_html = "".join(
    f'<a href="#cap_{i}" class="na" style="border-left-color:{CAP_COLORS.get(c,"#ccc")}">{he(c)}</a>'
    for i, c in enumerate(chapters.keys())
)

html = f"""<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8">
<title>5D Presupuesto — {he(MODEL_NAME)}</title>
<style>{css}</style></head>
<body>
<header>
  <div><h1>5D Quantity Takeoff — {he(MODEL_NAME)}</h1>
  <div style="font-size:11px;color:#9ab;margin-top:2px">{DATE_STR} · PyNET Platform</div></div>
  <div style="text-align:right">
    <div style="font-size:28px;font-weight:800;color:#f0f0f0">{fmt_eur(grand_total)}</div>
    <div style="font-size:11px;color:#9ab">Presupuesto Total</div>
  </div>
</header>
<div class="lay">
<nav>
  <div class="sb">
    <div class="sn">{len(budget)}</div>
    <div style="font-size:11px;color:#9ab;margin-top:2px">Partidas</div>
    <div class="ss2">{len(chapters)} capítulos · {total_elements} elementos</div>
  </div>
  <div class="nt">Capítulos</div>
  {nav_html}
</nav>
<main>
  <div class="kr">
    <div class="kp" style="border-color:#1a2332"><div class="v">{fmt_eur(grand_total)}</div><div class="l">Presupuesto Total</div></div>
    <div class="kp" style="border-color:#2980b9"><div class="v">{len(chapters)}</div><div class="l">Capítulos</div></div>
    <div class="kp" style="border-color:#27ae60"><div class="v">{len(budget)}</div><div class="l">Partidas</div></div>
    <div class="kp" style="border-color:#e67e22"><div class="v">{total_elements}</div><div class="l">Elementos medidos</div></div>
  </div>
  <div class="cr">
    <div class="cd"><h3>Distribución por Capítulo</h3>
      <div class="dw">{donut_svg()}<div class="lr">{legend}</div></div>
    </div>
    <div class="cd"><h3>Importe por Capítulo</h3>{bars_html}</div>
  </div>
  <div class="sg">{cap_cards}</div>
  {detail_html}
</main>
</div>
<footer>PyNET Platform · {he(MODEL_NAME)} · {DATE_STR} · Total: {fmt_eur(grand_total)}</footer>
</body></html>"""

html_path = Path(desktop) / f"5D_Presupuesto_{TODAY}.html"
with open(str(html_path), "w", encoding="utf-8") as f:
    f.write(html)

webbrowser.open(f'file:///{str(html_path).replace(chr(92), "/")}')
print(f"HTML: {html_path}")

ia_Result = {"excel": str(xls_path), "html": str(html_path), "grand_total": grand_total,
             "resources": len(budget), "chapters": dict(cap_totals)}
